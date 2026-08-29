"""Export experiment outputs to structured Excel workbooks."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def read_csv(path):

    with Path(path).open(
        "r",
        newline="",
        encoding="utf-8",
    ) as handle:

        return list(
            csv.DictReader(handle)
        )


def add_sheet(
    workbook,
    title,
    rows,
):

    worksheet = workbook.create_sheet(
        title=title[:31]
    )

    if not rows:
        return worksheet

    headers = list(
        rows[0].keys()
    )

    for column, header in enumerate(
        headers,
        start=1,
    ):

        cell = worksheet.cell(
            row=1,
            column=column,
            value=header,
        )

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for row_index, row in enumerate(
        rows,
        start=2,
    ):

        for column_index, header in enumerate(
            headers,
            start=1,
        ):

            value = row.get(
                header,
                "",
            )

            try:
                numeric = float(value)

                if (
                    numeric.is_integer()
                    and "." not in value
                ):
                    value = int(
                        numeric
                    )
                else:
                    value = numeric

            except (
                ValueError,
                TypeError,
            ):
                pass

            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for index in range(
        1,
        worksheet.max_column + 1,
    ):

        width = 12

        for cell in worksheet[
            get_column_letter(index)
        ]:

            if cell.value is not None:
                width = max(
                    width,
                    min(
                        len(
                            str(
                                cell.value
                            )
                        ) + 2,
                        40,
                    ),
                )

        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = width

    return worksheet


def export_results():

    metrics_dir = Path(
        "results/metrics"
    )

    metrics_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    default = workbook.active
    workbook.remove(default)

    matrix_csv = (
        metrics_dir
        / "full_experiment_matrix.csv"
    )

    aggregate_csv = (
        metrics_dir
        / "aggregated_results.csv"
    )

    if matrix_csv.exists():

        add_sheet(
            workbook,
            "Experiment Matrix",
            read_csv(matrix_csv),
        )

    if aggregate_csv.exists():

        add_sheet(
            workbook,
            "Aggregated Results",
            read_csv(aggregate_csv),
        )

    ablation_csv = (
        metrics_dir
        / "ablation_configurations.csv"
    )

    if ablation_csv.exists():

        add_sheet(
            workbook,
            "Ablation Config",
            read_csv(ablation_csv),
        )

    if not workbook.sheetnames:

        workbook.create_sheet(
            "README"
        )

        workbook["README"]["A1"] = (
            "Run the experiment and "
            "aggregation scripts first."
        )

    output = (
        metrics_dir
        / "adaptive_photonic_ml_results.xlsx"
    )

    workbook.save(
        output
    )

    print(
        f"Excel workbook saved to:\n"
        f"{output.resolve()}"
    )


if __name__ == "__main__":
    export_results()
