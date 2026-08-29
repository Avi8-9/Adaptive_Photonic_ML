"""Create structured Excel workbooks from actual experiment outputs."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def read_csv(path):

    with Path(path).open(
        "r",
        newline="",
        encoding="utf-8",
    ) as handle:

        return list(
            csv.DictReader(handle)
        )


def write_sheet(
    workbook,
    title,
    rows,
):

    worksheet = workbook.create_sheet(
        title=title[:31]
    )

    if not rows:
        return

    headers = list(
        rows[0].keys()
    )

    for column, header in enumerate(
        headers,
        start=1,
    ):

        cell = worksheet.cell(
            row=1,
            column=column,
            value=header,
        )

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for row_index, row in enumerate(
        rows,
        start=2,
    ):

        for column_index, header in enumerate(
            headers,
            start=1,
        ):

            value = row.get(
                header,
                "",
            )

            try:

                value = float(
                    value
                )

            except (
                ValueError,
                TypeError,
            ):
                pass

            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    worksheet.freeze_panes = "A2"

    if worksheet.max_row > 1:

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


def export_workbook(
    csv_paths,
    output_path,
):

    workbook = Workbook()

    workbook.remove(
        workbook.active
    )

    for csv_path in csv_paths:

        path = Path(csv_path)

        if not path.exists():
            continue

        rows = read_csv(path)

        write_sheet(
            workbook,
            path.stem,
            rows,
        )

    if not workbook.sheetnames:

        workbook.create_sheet(
            "No Results"
        )

    output_path = Path(
        output_path
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(
        output_path
    )

    print(
        "Workbook written:",
        output_path,
    )


if __name__ == "__main__":

    csv_files = list(
        Path("results")
        .rglob("*.csv")
    )

    export_workbook(
        csv_files,
        "results/tables/"
        "complete_experiment_archive.xlsx",
    )
